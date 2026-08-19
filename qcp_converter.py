"""
QCP → CNPE Excel 转换核心库（v3.7）
来源：skills/qcp-to-cnpe-excel/SKILL.md

提供以下函数：
- is_procedure_table(tbl) -> bool
- parse_qcp_pdf(pdf_path, supplier_count=1) -> List[dict]
  （v3.3 自动识别 A 版二重装备 / B 版 SEC-KSB）
- clean_steps_with_supplier_layers(steps, supplier_count=1) -> List[dict]
- _fix_template_styles(tmpl_path)  : 修复模板样式问题（openpyxl 需要）
- _open_template(tmpl_path)        : 加载并修复模板
- make_xlsx_0729(tmpl_path, out_path, steps, item_code, supplier_count=1) -> int
  （重写：用 openpyxl 加载、填数据、真正删除列）
- build_output_filename(pdf_path)  : 按 PDF 首页标题生成文件名（v3.6）
- extract_metadata_from_pdf(pdf_path) : 提取 19位编码/Rev/零件号/零件名（v3.6）

变更历史：
- v3.7 (2026-08-18) ：文件命名从 PDF 首页「题目/Title」提取，re.DOTALL 跨行处理。
- v3.6 (2026-08-18) ：见上文。
- v3.5 (2026-08-18) ：supplier_count 默认 None → 按 PDF 实际 a_point 自动判断，不再与 A/B 版挂钩。
- v3.4 (2026-08-18) ：补充 supplier_count=0 场景：B 版 SEC-KSB 无 A 列时
                A1/A2/A3 全删（不只删 A2/A3）。supplier_count<1 时也删 V 列（选点A1）。
                列删除逻辑：
                  0 = 删 T+U+V  → max_column=23
                  1 = 删 T+U    → max_column=24
                  2 = 删 T       → max_column=25
                  3 = 保留全部   → max_column=26
- v3.3 (2026-08-18) ：parse_qcp_pdf 自动识别 A 版（二重装备，col 11=A/14=S/16=C/18=O）
                与 B 版（SEC-KSB, col 9=S/12=C/15=O，无 A 列；remark col 19）。
                通过子表头「S」所在列索引自动选择列映射。
- v3.2 (2026-08-18) ：重构为 openpyxl 方案，真正删除 T/U 列（陈老师反馈）
                hidden 属性不能用了，会被上传系统检测到
- v3.1 (2026-08-17) ：hidden 属性方案（已废弃）
- v3.0 (2026-08-17) ：XML 手动删除列
- v2.1 (2026-08-17) ：A1/A2/A3 动态供应商分层逻辑

由 Peter 子Agent 调用，保持单一权威源。
"""


# ============================================================
# v3.7+ 自动同步检查
# 启动时验证 peter/SKILL.md 是否已同步本版本（耗时 < 1ms）
# ============================================================
__version__ = '3.7'

import os as _os
import re as _re
import warnings as _warnings


def _check_peter_sync():
    """启动时检查 peter/SKILL.md 是否已同步本版本

    - 不一致时输出 UserWarning（主Agent能捕获）
    - 异常时静默跳过（不阻塞转换）
    - 耗时 < 1ms（只读 400 字节）
    """
    try:
        peter_path = _os.path.join(
            _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
            'skills', 'peter', 'SKILL.md'
        )
        if not _os.path.exists(peter_path):
            return
        with open(peter_path, 'r', encoding='utf-8') as _f:
            head = _f.read(400)
        # peter/SKILL.md 的 version_note 字段应包含「qcp_converter.py vX.Y」
        m = _re.search(r'qcp_converter\.py\s*v(\d+\.\d+)', head)
        if m and m.group(1) != __version__:
            _warnings.warn(
                f'\n'
                f'⚠️  qcp_converter.py v{__version__} 已升版，'
                f'但 peter/SKILL.md 还在引用 v{m.group(1)}。\n'
                f'请同步：1) peter 的 version/version_note  '
                f'2) peter 调用示例  3) qcp-to-cnpe-excel/SKILL.md\n'
                f'（参考：见 .learnings/LEARNINGS.md "技能文档也要同步更新"）',
                stacklevel=2,
            )
    except Exception:
        pass  # 不阻塞转换


# 模块加载时自动跑一次（仅一次，spawn 一次只触发一次）
_check_peter_sync()


import os, shutil, re, zipfile
import pdfplumber


# ============================================================
# 表格过滤
# ============================================================
def is_procedure_table(tbl):
    """判断是否为工序表（非文件清单表）"""
    if not tbl or len(tbl) < 3:
        return False
    # 收集整个表的所有文本
    full_text = ''
    for row in tbl:
        for c in row:
            full_text += str(c or '') + ' '
    # 工序表特征：含「工序名称」+「作业依据文件」+ A/S/C/O 选点标签
    if '工序名称' in full_text and ('作业依据文件' in full_text or 'Process' in full_text):
        # 子表头包含 A/S/C/O 选点字母（独立 cell）
        has_point_cols = False
        for row in tbl:
            for c in row:
                cell = str(c or '').strip()
                if cell in ('A', 'S', 'C', 'O'):
                    has_point_cols = True
                    break
            if has_point_cols:
                break
        return has_point_cols
    return False


# ============================================================
# PDF 解析
# ============================================================
def _detect_column_mapping(tbl, header_idx):
    """根据子表头中「S」所在列索引自动判断 A/B 版本列映射

    返回 dict 包含 a_col/s_col/c_col/o_col/remark_col
    如果识别失败则抛 ValueError。

    A 版（二重装备）：A=col11, S=col14, C=col16, O=col18, remark=col22
    B 版（SEC-KSB）  ：A=不存在, S=col9, C=col12, O=col15, remark=col19
    """
    if header_idx < 0 or header_idx + 1 >= len(tbl):
        raise ValueError('header_idx 越界，无法定位子表头')
    sub = tbl[header_idx + 1]

    # 找到 S/C/O 标记的列索引（cell 内容必须严格等于 'S' / 'C' / 'O'）
    def find_marker(label):
        for j, c in enumerate(sub):
            if c and str(c).strip() == label:
                return j
        return None

    s_col = find_marker('S')
    c_col = find_marker('C')
    o_col = find_marker('O')
    a_col = find_marker('A')  # B 版通常没有

    if s_col is None or c_col is None or o_col is None:
        raise ValueError(
            f'工序表子表头缺 S/C/O 标记（found S={s_col} C={c_col} O={o_col}）'
        )

    # remark 列：定位到「备注」所在列（A/B 版均在主表头 row[header_idx]）
    remark_col = None
    for j, c in enumerate(tbl[header_idx]):
        if c and '备注' in str(c):
            remark_col = j
            break

    return {
        'a_col': a_col,         # None if not present (B 版)
        's_col': s_col,
        'c_col': c_col,
        'o_col': o_col,
        'remark_col': remark_col,
        # 版本标识（仅供调试/报告使用）
        'version': 'B' if a_col is None else 'A',
    }


def _read_cell(row, col):
    """读取 row[col]，空值返回 ''"""
    if col is None or col >= len(row):
        return ''
    v = row[col]
    if not v:
        return ''
    return str(v).strip()


def parse_qcp_pdf(pdf_path, supplier_count=None):
    """
    解析 QCP PDF，提取工序数据（v3.5 自动识别 A/B 版 + supplier_count）

    陈老师 2026-08-18 反馈：供应商多少跟 A/B 版没有直接关系，
    需要根据质量计划实际情况进行动态调整。

    supplier_count 参数（v3.5 改为可选）：
    - None（默认）：自动推断
      - PDF 有 A 列且 a_point 有值 → supplier_count = 1
      - PDF 无 A 列 → supplier_count = 0
      - PDF 有 A 列但 a_point 全空 → supplier_count = 0
    - 0/1/2/3：强制指定（默认 =1）

    自动识别列映射：
    - A 版二重装备（横向 841×595）：A=col11, S=col14, C=col16, O=col18, remark=col22
    - B 版 SEC-KSB  （竖版 596×842）：S=col9, C=col12, O=col15, remark=col19, 无 A 列
    """
    all_steps = []
    col_map = None  # 延迟到第一个工序表识别后填入
    with pdfplumber.open(pdf_path) as pdf:
        for pg_idx in range(len(pdf.pages)):
            page = pdf.pages[pg_idx]
            tables = page.extract_tables()
            for tbl in tables:
                if not is_procedure_table(tbl):
                    continue
                header_idx = -1
                for ri, row in enumerate(tbl):
                    if row and '序号' in str(row[0] or ''):
                        header_idx = ri
                        break
                if header_idx < 0:
                    continue

                # 第一次见到工序表时识别列映射；之后所有页都按同一映射解析
                if col_map is None:
                    try:
                        col_map = _detect_column_mapping(tbl, header_idx)
                    except ValueError as e:
                        # 跳过无法识别的工序表
                        continue
                s_col = col_map['s_col']
                c_col = col_map['c_col']
                o_col = col_map['o_col']
                a_col = col_map['a_col']  # 可能为 None
                remark_col = col_map['remark_col']

                # 从 header_idx+2 起解析数据
                for ri in range(header_idx + 2, len(tbl)):
                    row = tbl[ri]
                    if not row or not row[0]:
                        continue
                    proc_no = str(row[0]).strip()
                    if not proc_no or proc_no in ['.', '', '序号']:
                        continue
                    name = _read_cell(row, 1)
                    if not name or name == '.':
                        continue
                    name = re.sub(r'\n', ' ', name)

                    a_val = _read_cell(row, a_col) if a_col is not None else ''
                    s_val = _read_cell(row, s_col)
                    c_val = _read_cell(row, c_col)
                    o_val = _read_cell(row, o_col)
                    remark = _read_cell(row, remark_col) if remark_col is not None else ''
                    remark = re.sub(r'\n', ' ', remark)

                    all_steps.append({
                        'proc_no': proc_no,
                        'name': name,
                        'a_point': a_val if a_val in ('H', 'R', 'W') else '',
                        's_point': s_val if s_val in ('H', 'R', 'W') else '',
                        'c_point': c_val if c_val in ('H', 'R', 'W') else '',
                        'o_point': o_val if o_val in ('H', 'R', 'W') else '',
                        'remark': remark if remark not in ('.', '') else ''
                    })
    return all_steps


def _auto_detect_supplier_count(steps, col_map):
    """根据 PDF 实际数据自动判断 supplier_count

    陈老师 2026-08-18 反馈：供应商多少跟 A/B 版没有直接关系，需要动态调整。

    规则：
    - PDF 有 A 列（col_map['a_col'] is not None）
        - 工序中实际有 a_point 值（H/R/W） → supplier_count = 1
        - 工序中全为空 → supplier_count = 0（PDF 有 A 列但实际未使用）
    - PDF 无 A 列（col_map['a_col'] is None，B 版） → supplier_count = 0
    """
    has_a_col = col_map.get('a_col') is not None
    if not has_a_col:
        return 0
    # 检查工序中是否有实际 a_point
    any_a = any(s.get('a_point') for s in steps)
    return 1 if any_a else 0


def get_detected_version(pdf_path):
    """返回该 PDF 识别出的 QCP 版本（'A' / 'B'），未识别返回 None。

    仅用于报告/日志，不影响 parse_qcp_pdf 自身行为。
    """
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                if not is_procedure_table(tbl):
                    continue
                header_idx = -1
                for ri, row in enumerate(tbl):
                    if row and '序号' in str(row[0] or ''):
                        header_idx = ri
                        break
                if header_idx < 0:
                    continue
                try:
                    cm = _detect_column_mapping(tbl, header_idx)
                    return cm['version']
                except ValueError:
                    continue
    return None


# ============================================================
# 清洗与供应商分层
# ============================================================
def clean_steps_with_supplier_layers(steps, supplier_count=1):
    """
    清洗数据，按供应商数量填 A1/A2/A3
    supplier_count: 1=只填A1, 2=A1+A2, 3=全填
    """
    result = []
    for s in steps:
        s_display = s['s_point'] + '点' if s['s_point'] in ('H', 'W', 'R') else None
        a_display = s['a_point'] + '点' if s['a_point'] in ('H', 'W', 'R') else None
        c_display = s['c_point'] + '点' if s['c_point'] in ('H', 'W', 'R') else None
        o_display = s['o_point'] + '点' if s['o_point'] in ('H', 'W', 'R') else None

        # 供应商分层（按 supplier_count 决定填几层）
        # 注：当前实现假设只有 1 个供应商的 A1 列数据
        # 如果有 A2/A3，需要从 PDF 不同列位提取
        a1_display = a_display if supplier_count >= 1 else None
        a2_display = None  # TODO: A2 数据提取（如有）
        a3_display = None  # TODO: A3 数据提取（如有）

        has_sp = s['s_point'] in ('H', 'W', 'R')
        is_precheck = '先决条件检查' in s['name']
        is_close = '质量计划关闭' in s['name'] or '计划关闭' in s['name']
        q_val = 'N' if (is_precheck or is_close) else ('Y' if has_sp else 'N')

        result.append({
            'proc_no': s['proc_no'],
            'name': s['name'],
            's_point': s_display,
            'c_point': c_display,
            'o_point': o_display,
            'a1_point': a1_display,
            'a2_point': a2_display,
            'a3_point': a3_display,
            'q': q_val,
            'remark': s.get('remark', '')
        })
    return result


# ============================================================
# 列删除（A2/A3 未用时）
# ============================================================
def delete_columns(sheet_path, cols_to_delete):
    """从 xlsx 中真正删除指定列

    4 步处理（陈老师 2026-08-18 反馈）：
    - ❌ 不要用 hidden 属性！上传系统会检测到列还在 → 失败
    - ✅ 必须真正删除：cell 引用、cols 定义、dataValidation 全部清理
    - ✅ Excel 删除后会自动让后面的列顶位（V 列顶位变成 T）

    1. <cols> 定义：删除该列的列宽定义
    2. <c> 引用：从所有行中移除该列的 cell 引用
    3. <dataValidation>：删除指向该列的数据验证规则
    4. <dataValidations count="N">：修正 count 值
    """
    with open(sheet_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 1. 从 <cols> 定义中删除该列
    for col in cols_to_delete:
        col_idx = ord(col) - 64  # A=1, B=2, ..., T=20, U=21, V=22
        content = re.sub(
            rf'<col min="{col_idx}" max="{col_idx}"[^/]*/>',
            '',
            content
        )

    # 2. 从所有行中移除该列的 <c> 引用（支持 inline string 和 v 两种格式）
    for col in cols_to_delete:
        content = re.sub(
            rf'<c r="{col}\d+"[^/]*(?:/>|.*?</c>)',
            '',
            content,
            flags=re.DOTALL
        )

    # 3. ⚠️ 删除指向该列的 <dataValidation> 规则
    for col in cols_to_delete:
        # 单个 <dataValidation ...><formula1>...</formula1></dataValidation>
        content = re.sub(
            rf'<dataValidation[^>]*sqref="{col}\d+:{col}\d+"[^>]*>.*?</dataValidation>',
            '',
            content,
            flags=re.DOTALL
        )
        # 自闭合 <dataValidation ... sqref="..." />
        content = re.sub(
            rf'<dataValidation[^/]*sqref="{col}\d+:{col}\d+"[^/]*/>',
            '',
            content
        )

    # 4. 修正 <dataValidations count="N"> 的 count 值
    actual = len(re.findall(r'<dataValidation', content))
    content = re.sub(
        r'<dataValidations count="\d+">',
        f'<dataValidations count="{actual}">',
        content,
        count=1
    )

    # ⚠️ 不添加 hidden 属性！隐藏列会被上传系统检测到，仍然算存在该列

    with open(sheet_path, 'w', encoding='utf-8') as f:
        f.write(content)


# ============================================================
# Excel 生成（0729 新模板）
# ============================================================
def make_xlsx_0729(tmpl_path, out_path, steps, item_code, supplier_count=1):
    """
    按 0729 新模板生成 Excel，并按供应商数量真正删除 A2/A3 列

    v3.2 重写：用 openpyxl 加载、填数据、delete_cols API
    （不用 XML hack，不用 hidden 属性——上传系统会检测到 hidden 列）

    模板列字母定义（实测验证 v3.2，2026-08-18）：
      A 错误信息 | B 过滤标识 | C *排序号 | D 工序唯一标识码
      E *物项标识码 | F 标准工序编码 | G *工序编号 | H *工序名称
      I 工序描述 | J 供应商选点 | K cnpe选点 | L 业主选点
      M 前置工序 | N 依据文件信息 | O 厂家内部依据文件编号 | P 厂家内部依据文件名称
      Q 厂家内部依据文件版本 | R 是否可选执行 | S 是否产生报告
      **T 选点A3** | **U 选点A2** | **V 选点A1**
      W *实物是否须有编号 | X 监管单位选点类型 | Y 备注 | Z id

    供应商场景（陈老师 2026-08-18 反馈）：
      supplier_count=1：删除 T(U=选点A2 顶位变成 T) + U 删后 V=选点A1 顶位变成 U
                      → 原 W 顶位变成 V, 原 X 顶位变成 W, 原 Y 顶位变成 X
      supplier_count=2：删除 T(U=选点A2 顶位变成 T)
                      → 原 V=选点A1 不动, 原 W/Y/X 顶位上移
      supplier_count=3：保留所有列

    重要：必须真正删除列（不要用 hidden 属性！上传系统会检测到 hidden 列失败）
    """
    from openpyxl import load_workbook

    # 1. 修复模板的 styles.xml 问题（named_styles.name 为 None）
    fixed_tmpl = _fix_template_styles(tmpl_path)

    # 2. 加载模板
    wb = load_workbook(fixed_tmpl)
    ws = wb['质量计划-工序信息']

    # 3. 填数据到模板（Row 10 起）
    # 实际列定义（实测验证 v3.2）：
    #   C=*排序号 | E=*物项标识码 | G=*工序编号 | H=*工序名称
    #   J=供应商选点 | K=cnpe选点 | L=业主选点
    #   S=是否产生报告 | T=选点A3 | U=选点A2 | V=选点A1
    #   W=*实物是否须有编号 | X=监管单位选点类型 | Y=备注
    for i, step in enumerate(steps):
        row_num = 10 + i
        # C 列：排序号
        ws.cell(row=row_num, column=3, value=10 + i * 10)
        # E 列：物项标识码
        ws.cell(row=row_num, column=5, value=item_code)
        # G 列：工序编号
        ws.cell(row=row_num, column=7, value=step['proc_no'])
        # H 列：工序名称
        ws.cell(row=row_num, column=8, value=step['name'])
        # J 列：供应商选点
        if step.get('s_point'):
            ws.cell(row=row_num, column=10, value=step['s_point'])
        # K 列：cnpe选点
        if step.get('c_point'):
            ws.cell(row=row_num, column=11, value=step['c_point'])
        # L 列：业主选点
        if step.get('o_point'):
            ws.cell(row=row_num, column=12, value=step['o_point'])
        # S 列：是否产生报告
        ws.cell(row=row_num, column=19, value=step['q'])
        # T 列：选点A3（3 家供应商才填）
        if step.get('a3_point') and supplier_count >= 3:
            ws.cell(row=row_num, column=20, value=step['a3_point'])
        # U 列：选点A2（2 家及以上供应商才填）
        if step.get('a2_point') and supplier_count >= 2:
            ws.cell(row=row_num, column=21, value=step['a2_point'])
        # V 列：选点A1（始终填）
        if step.get('a1_point'):
            ws.cell(row=row_num, column=22, value=step['a1_point'])
        # W 列：*实物是否须有编号
        ws.cell(row=row_num, column=23, value='否')
        # Y 列：备注（有内容才填）
        if step.get('remark'):
            ws.cell(row=row_num, column=25, value=step['remark'])

    # 4. ⭐ 真正删除未用的 A1/A2/A3 列（用 openpyxl delete_cols API）
    # 陈老师 2026-08-18 反馈：当 A1/A2/A3 都不存在时需要全部删除
    # supplier_count 含义：
    #   0 = 没有 A 选点（B 版 SEC-KSB：只有 S/C/O）→ 删 T/U/V 三列
    #   1 = 只填 A1（单供应商） → 删 T/U 两列（保留 V）
    #   2 = 填 A1+A2 → 删 T 一列
    #   3 = 填 A1+A2+A3 → 三列都保留
    # 从右到左删（openpyxl 不会自动调整索引）：
    if supplier_count < 3:
        ws.delete_cols(20)  # 删 T 列（选点A3）
    if supplier_count < 2:
        ws.delete_cols(20)  # 删 U 列（选点A2，删 T 后变成第 20 列）
    if supplier_count < 1:
        ws.delete_cols(20)  # 删 V 列（选点A1，删 T+U 后变成第 20 列）

    # 5. 保存
    wb.save(out_path)
    return len(steps)


def _fix_template_styles(tmpl_path):
    """修复模板的 styles.xml 问题（named_styles.name 为 None）

    openpyxl 加载某些模板会报错：cellStyle.name should be str but value is NoneType
    修复方法：将 <cellStyle xfId="N"/> 改成 <cellStyle name="Default" xfId="N"/>
    """
    import tempfile
    work = tempfile.mkdtemp(prefix='tmpl_fix_')
    try:
        with zipfile.ZipFile(tmpl_path) as z:
            z.extractall(work)
        styles_path = os.path.join(work, 'xl', 'styles.xml')
        with open(styles_path) as f:
            content = f.read()
        # 修复 1：name='' 的 cellStyle
        new_content = re.sub(r'<cellStyle name=""', '<cellStyle name="Default"', content)
        # 修复 2：name 缺失的 cellStyle（<cellStyle xfId="N"/>）
        new_content = re.sub(
            r'<cellStyle xfId="(\d+)"(\s*/>)',
            r'<cellStyle name="Default" xfId="\1"\2',
            new_content
        )
        if new_content != content:
            with open(styles_path, 'w') as f:
                f.write(new_content)
            # 重新打包
            fixed_path = tmpl_path + '.fixed.xlsx'
            if os.path.exists(fixed_path):
                os.remove(fixed_path)
            with zipfile.ZipFile(fixed_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for root, dirs, files in os.walk(work):
                    for file in files:
                        fp = os.path.join(root, file)
                        zout.write(fp, os.path.relpath(fp, work))
            return fixed_path
        else:
            return tmpl_path  # 无需修复
    finally:
        import shutil as _sh
        _sh.rmtree(work, ignore_errors=True)


# ============================================================
# 便捷入口（高级封装）
# ============================================================
def convert_qcp_to_cnpe(pdf_path, tmpl_path, item_code='1907RCP10101', supplier_count=None,
                       out_path=None):
    """
    一站式转换：PDF → Excel

    v3.5 陈老师反馈 2026-08-18：
    - supplier_count=None 时根据 PDF 实际数据自动判断（不是与 A/B 版挂钩）
    - 文件名按 19 位编码 + Rev 版本 + 零件号 + 零件名 生成

    Returns: (out_path, num_procedures)
    """
    import tempfile
    if out_path is None:
        out_path = os.path.join(
            tempfile.gettempdir(),
            f'qcp_cnpe_{os.path.basename(pdf_path).replace(".pdf","")}_{os.getpid()}.xlsx'
        )

    # 泵壳及子件特殊规则（2026-08-19 陈老师反馈）：
    # 泵壳本体及所有泵壳相关子件（底脚/安全端/焊接见证件/母材见证件/焊材）
    # 共用零件号 101.01 → 物项标识码固定为 1907RCP10101
    # 识别方法：item_code 为 19 位编码时，Z 段为 Z4x/Z5x/Z6x/Z7x → 强制覆盖
    if len(item_code) >= 11 and item_code[8:11].startswith(('Z4', 'Z5', 'Z6', 'Z7')):
        print(f'[info] 泵壳相关 QCP (Z段={item_code[8:11]}) → 强制使用 1907RCP10101')
        item_code = '1907RCP10101'

    steps = parse_qcp_pdf(pdf_path, supplier_count=supplier_count)
    # 自动推断 supplier_count（如果未指定）
    if supplier_count is None:
        col_map = get_column_map(pdf_path) or {}
        supplier_count = _auto_detect_supplier_count(steps, col_map)
    cleaned = clean_steps_with_supplier_layers(steps, supplier_count=supplier_count)
    n = make_xlsx_0729(tmpl_path, out_path, cleaned, item_code, supplier_count=supplier_count)
    return out_path, n, supplier_count


def get_column_map(pdf_path):
    """获取 PDF 的列映射（供其他地方调用）"""
    import pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                if not is_procedure_table(tbl):
                    continue
                header_idx = -1
                for ri, row in enumerate(tbl):
                    if row and '序号' in str(row[0] or ''):
                        header_idx = ri
                        break
                if header_idx < 0:
                    continue
                try:
                    return _detect_column_mapping(tbl, header_idx)
                except ValueError:
                    continue
    return None


def extract_metadata_from_pdf(pdf_path):
    """从 PDF 中提取 19 位编码、Rev 版本、零件号、零件名

    陈老师 2026-08-18 反馈：
    - 文件命名需要根据文件的19位编码信息命名
    - 同时加上 Rev.版本
    - 首页标题才是真正的标题（不是部件信息表的零件名称）

    策略：
    - 19 位编码、Rev：从文件名正则提取（最可靠）
    - 零件名、零件号：从 PDF 首页的「题目/Title」栏提取

    返回 dict: {'code19': 'SMX44400Z69101A04GN', 'rev': 'B', 'part_no': '230.11', 'part_name': '叶轮（230.11）'}
    """
    import re
    import pdfplumber
    fname = os.path.basename(pdf_path)
    meta = {'code19': '', 'rev': '', 'part_no': '', 'part_name': ''}

    # 1. 从 PDF 文件名提取 19 位编码（最可靠）
    m19 = re.search(r'SMX\d{5}Z\d{2}101A04GN', fname)
    if m19:
        meta['code19'] = m19.group()

    # 2. 从 PDF 文件名提取 Rev 版本
    mrev = re.search(r'Rev\.?([A-Z]\d*)', fname, re.IGNORECASE)
    if mrev:
        meta['rev'] = mrev.group(1)

    # 3. 从 PDF 首页文本提取题目/Title（陈老师 2026-08-18 反馈：命名按首页标题）
    with pdfplumber.open(pdf_path) as pdf:
        first_page_text = pdf.pages[0].extract_text() or ''

    # 提取题目/Title 后的内容
    # 例如：「题目/Title:\n反应堆冷却剂泵叶轮（230.11）成品制造质量计划\nQuality Plan ...」
    # 或：「题目/Title：反应堆冷却剂泵底脚与泵壳组焊焊接见证件制造质量计划」
    # 使用 re.DOTALL 以处理 Title 后面跨多行的中文标题
    m_title = re.search(r'题目\s*[/:：]?\s*Title\s*[:：]?\s*\n?(.+?)(?:Quality Plan|$)', first_page_text, re.DOTALL)
    if m_title:
        # Title 行可能是中文（陈老师报告名称）或英文（Quality Plan ...）
        title_cn = m_title.group(1).strip()
        # 题目格式：「反应堆冷却剂泵<零件名>质量计划」或「反应堆冷却剂泵<零件名>成品制造质量计划」
        # 提取「泵」到「质量计划」之间的零件名
        m_part = re.search(r'泵(.+?)\s*质量计划', title_cn, re.DOTALL)
        if m_part:
            part_raw = m_part.group(1).strip()
            # 去掉「成品制造」/「制造」后缀（如果有）
            part_raw = re.sub(r'(?:成品)?(?:制造|加工|复验|检验|试验)$', '', part_raw)
            # 单独提取零件号
            m_partno = re.search(r'[（(](\d{3}\.\d{2})[)）]', part_raw)
            if m_partno:
                meta['part_no'] = m_partno.group(1)
            # 零件名：去掉括号及括号内零件号，只留中文名
            # 如「叶轮（230.11）」→ 「叶轮」
            part_name = re.sub(r'[（(][^）)]*[）)]', '', part_raw).strip()
            meta['part_name'] = part_name

    return meta


def build_output_filename(pdf_path, version='v14', suffix='_new'):
    """根据 19 位编码、Rev、零件号、零件名生成输出文件名

    格式：{19位编码}_Rev.{版本}_{零件号}_{零件名}_CNPE_{version}{suffix}.xlsx
    例：SMX44400Z69101A04GN_Rev.B_230.11_叶轮_CNPE_v14_new.xlsx

    陈老师 2026-08-18 反馈：文件名要加 Rev.版本
    """
    meta = extract_metadata_from_pdf(pdf_path)
    parts = []
    if meta['code19']:
        parts.append(meta['code19'])
    if meta['rev']:
        parts.append(f"Rev.{meta['rev']}")
    if meta['part_no']:
        parts.append(meta['part_no'])
    if meta['part_name']:
        parts.append(meta['part_name'])
    parts.append(f'CNPE_{version}')
    if suffix:
        parts.append(suffix.lstrip('_'))
    return '_'.join(parts) + '.xlsx'


if __name__ == '__main__':
    # 自测
    import sys
    if len(sys.argv) >= 3:
        pdf = sys.argv[1]
        tmpl = sys.argv[2]
        item = sys.argv[3] if len(sys.argv) > 3 else '1907RCP10101'
        sc = int(sys.argv[4]) if len(sys.argv) > 4 else 1
        out, n = convert_qcp_to_cnpe(pdf, tmpl, item, sc)
        print(f'✅ {n} 道工序 → {out}')
    else:
        print('用法: python qcp_converter.py <pdf> <tmpl.xlsx> [item_code] [supplier_count]')