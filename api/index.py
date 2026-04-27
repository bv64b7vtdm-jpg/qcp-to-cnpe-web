"""
Flask web app for QCP to CNPE conversion - generic version.
Parses any QCP PDF table using pdfplumber.
"""

import os
import uuid
import re
import pdfplumber
from flask import Flask, request, render_template, send_file, jsonify
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.secret_key = os.environ.get('SECRET_KEY', 'qcp-cnpe-secret-key-2026')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = '/tmp/qcp-uploads'

ALLOWED_EXTENSIONS = {'pdf'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/health')
def health():
    return jsonify({'status': 'ok'})


@app.route('/upload', methods=['POST'])
def upload():
    try:
        item_code_19 = request.form.get('item_code_19', '').strip()
        supplier_item_code = request.form.get('supplier_item_code', '').strip()
        part_no = request.form.get('part_no', '').strip()

        if not item_code_19 or not supplier_item_code:
            return jsonify({'error': '19位编码和厂家物项编号为必填项'}), 400

        file = request.files.get('file')
        if not file or file.filename == '':
            return jsonify({'error': '请上传PDF文件'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': '只支持PDF格式文件'}), 400

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{uuid.uuid4().hex}_{secure_filename(file.filename)}")
        file.save(pdf_path)

        # Parse QCP data from PDF
        qcp_data = extract_qcp_data(pdf_path)

        if not qcp_data:
            os.remove(pdf_path)
            return jsonify({'error': '未能从PDF中识别到QCP工序数据。请确认PDF为文本型（不是扫描件），且包含工序号列表。'}), 422

        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"CNPE_转换_{item_code_19}.xlsx")
        template_path = os.path.join(BASE_DIR, 'templates', 'CNPE_质量计划导入Excel模板.xlsx')

        if os.path.exists(template_path):
            from utils.excel_filler import fill_cnpe_template
            fill_cnpe_template(template_path, output_path, qcp_data, item_code_19, part_no, supplier_item_code)
        else:
            # No template - generate from scratch
            generate_cnpe_excel(output_path, qcp_data, item_code_19, part_no, supplier_item_code)

        os.remove(pdf_path)

        return send_file(output_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"CNPE_转换_{item_code_19}.xlsx")

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def extract_qcp_data(pdf_path):
    """Extract QCP steps from any QCP PDF using table detection."""
    all_steps = []
    seen_keys = set()

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 2:
                        continue

                    # Try multiple column positions for step number
                    cell0 = (row[0] or '').strip()
                    cell1 = (row[1] or '').strip()

                    # Find the step number column (letter + number pattern like A1, B2.3, C1.1.2)
                    step_num = None
                    step_name = None

                    # Check cell0 as step number
                    if cell0 and re.match(r'^[A-Z]\d+', cell0):
                        step_num = cell0
                        step_name = cell1
                    # Check cell1 as step number (name in cell0)
                    elif cell1 and re.match(r'^[A-Z]\d+', cell1):
                        step_num = cell1
                        step_name = cell0

                    if step_num and step_name:
                        # Clean up step name
                        name_lines = step_name.split('\n')
                        clean_name = name_lines[0].strip() if name_lines else step_name.strip()

                        # Skip if name is too short or looks like a header
                        if len(clean_name) < 2:
                            continue
                        if clean_name in ('序号', 'No.', '工序名称', 'Process Name', '报告编号', 'Remark'):
                            continue

                        # Skip pure English without Chinese (might be headers)
                        has_chinese = re.search(r'[\u4e00-\u9fa5]', clean_name)
                        has_meaningful_content = len(clean_name) >= 2
                        if not (has_chinese or has_meaningful_content):
                            continue

                        key = f"{step_num}|{clean_name}"
                        if key not in seen_keys:
                            seen_keys.add(key)
                            all_steps.append({
                                'step': step_num,
                                'name': clean_name,
                                'content': ' '.join(name_lines[1:]).strip() if len(name_lines) > 1 else ''
                            })
                        continue

                    # Try cell1 alone
                    if cell1 and re.match(r'^[A-Z]\d+(?:\.\d+)*$', cell1.strip()):
                        step_num = cell1.strip()
                        step_name = cell0
                        if step_name and len(step_name) >= 2:
                            name_lines = step_name.split('\n')
                            clean_name = name_lines[0].strip()
                            key = f"{step_num}|{clean_name}"
                            if key not in seen_keys:
                                seen_keys.add(key)
                                all_steps.append({
                                    'step': step_num,
                                    'name': clean_name,
                                    'content': ''
                                })

    # Sort by step number
    def sort_key(s):
        num = re.sub(r'[A-Z]', '', s['step'])
        parts = num.split('.')
        return [int(p) if p.isdigit() else 0 for p in parts]

    all_steps.sort(key=sort_key)
    return all_steps


def generate_cnpe_excel(output_path, qcp_data, item_code_19, part_no, supplier_item_code):
    """Generate CNPE Excel from QCP data without template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        # Sheet 1: QCP工序
        ws = wb.active
        ws.title = 'QCP工序'

        # Header
        ws['A1'] = 'QCP转CNPE质量计划'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws['A3'] = '19位编码'
        ws['B3'] = item_code_19
        ws['A4'] = '零件号'
        ws['B4'] = part_no or ''
        ws['A5'] = '厂家物项编号'
        ws['B5'] = supplier_item_code
        ws['A6'] = '生成日期'
        from datetime import datetime
        ws['B6'] = datetime.now().strftime('%Y-%m-%d')

        # Table header
        headers = ['序号', '工序号', '工序名称', '质量控制点', '报告编号', '备注']
        header_row = 8
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for i, row_data in enumerate(qcp_data):
            row_num = header_row + 1 + i
            ws.cell(row=row_num, column=1, value=i + 1)
            ws.cell(row=row_num, column=2, value=row_data['step'])
            ws.cell(row=row_num, column=3, value=row_data['name'])
            ws.cell(row=row_num, column=4, value='')
            ws.cell(row=row_num, column=5, value='')
            ws.cell(row=row_num, column=6, value=row_data.get('content', ''))

            # Alternate row colors
            if i % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color='F8F6FF', end_color='F8F6FF', fill_type='solid')

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30

        # Sheet 2: 填写说明
        ws2 = wb.create_sheet('填写说明')
        notes = [
            ['QCP转CNPE填写说明'],
            [''],
            ['字段', '说明', '示例'],
            ['19位编码', '采购订单中的19位物料编码', '1P12345678901234567'],
            ['零件号', '可选，零件编号', 'A19'],
            ['厂家物项编号', '供应商提供的物项编号', 'SK93303-B1'],
            ['序号', '自动生成的序号', '1, 2, 3...'],
            ['工序号', 'QCP中的编号', 'B1.0, B1.1, C1.1.1...'],
            ['工序名称', '该工序的名称', '水压试验、性能试验...'],
            ['质量控制点', 'R/W/H', 'W'],
            ['报告编号', '对应的检验程序编号', 'SKR80004540...'],
            ['备注', '检验内容或备注', ''],
        ]
        for row_idx, row_data in enumerate(notes, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:
                    cell.font = Font(bold=True)
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 30

        wb.save(output_path)

    except ImportError:
        # Fallback: use XLSXWriter equivalent via openpyxl or simple csv
        import csv
        csv_path = output_path.replace('.xlsx', '.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['QCP转CNPE质量计划'])
            writer.writerow(['19位编码', item_code_19])
            writer.writerow(['零件号', part_no or ''])
            writer.writerow(['厂家物项编号', supplier_item_code])
            writer.writerow(['生成日期', datetime.now().strftime('%Y-%m-%d')])
            writer.writerow([])
            writer.writerow(['序号', '工序号', '工序名称', '备注'])
            for i, row_data in enumerate(qcp_data, 1):
                writer.writerow([i, row_data['step'], row_data['name'], row_data.get('content', '')])
        # Convert CSV to XLSX using openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'QCP工序'
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        wb.save(output_path)
        os.remove(csv_path)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
