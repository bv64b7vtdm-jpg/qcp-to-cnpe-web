#!/usr/bin/env python3
"""
转换 Z621（底脚组焊）和 Z631（安全端焊缝）QCP PDF → CNPE Excel
工序编号格式：严格按PDF原始格式，纯数字（1, 2, 3, 4-1, 5-2）不加C后缀
"""
import sys
import os
import re
import shutil
from collections import defaultdict

sys.path.insert(0, '/home/songsong/.openclaw/workspace/qcp-to-cnpe-web/')
import pdfplumber
from openpyxl import load_workbook

OUTPUT_DIR = "/mnt/d/SynologyDrive/QCP to Excel/"
TEMPLATE_FILE = "/home/songsong/.openclaw/workspace/qcp-to-cnpe-web/template/CNPE_质量计划导入Excel模板.xlsx"

COL_A_X = (485, 495)
COL_S_X = (550, 560)
COL_C_X = (618, 628)
COL_O_X = (685, 695)


def extract_col(chars, x_range):
    matching = [c for c in chars if x_range[0] <= c['x0'] <= x_range[1]]
    return ''.join(c['text'] for c in sorted(matching, key=lambda x: x['x0'])).strip()


def get_row_text(chars):
    return ''.join(c['text'] for c in sorted(chars, key=lambda x: x['x0']) if c['text'].strip())


def parse_qcp(pdf_path):
    """解析QCP PDF，返回工序列表"""
    results = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            chars = page.chars
            rows = defaultdict(list)
            for c in chars:
                if c['text'].strip():
                    rows[round(c['top'])].append(c)
            y_keys = sorted(rows.keys())
            
            # Build ASCO lookup for this page
            asco_lookup = {}
            for y in y_keys:
                row_chars = rows[y]
                a = extract_col(row_chars, COL_A_X)
                s = extract_col(row_chars, COL_S_X)
                c_col = extract_col(row_chars, COL_C_X)
                o = extract_col(row_chars, COL_O_X)
                if a in ('H', 'W', 'R') or s in ('H', 'W', 'R') or c_col in ('H', 'W', 'R') or o in ('H', 'W', 'R'):
                    asco_lookup[y] = (a, s, c_col, o)
            
            page_asco_ys = sorted(asco_lookup.keys())
            
            # Find proc rows (digit at x<60)
            for y_idx, y in enumerate(y_keys):
                row_chars = rows[y]
                sorted_p = sorted(row_chars, key=lambda x: x['x0'])
                
                if not (sorted_p and sorted_p[0]['text'].isdigit() and sorted_p[0]['x0'] < 60):
                    continue
                
                # Extract proc number: only look at chars with x < 60
                # (proc number is always at leftmost part of row)
                all_digits = [c for c in sorted_p if c['x0'] < 60 and c['text'].isdigit()]
                if not all_digits:
                    continue
                
                first_digit = all_digits[0]
                
                # Collect all digit and hyphen chars from same row (y within 3) at x < 60
                digit_hyphen = sorted(
                    [c for c in row_chars 
                     if c['text'] in ('-', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9')
                     and c['x0'] < 60
                     and abs(c['top'] - first_digit['top']) < 3],
                    key=lambda x: x['x0']
                )
                proc_no = ''.join(c['text'] for c in digit_hyphen)
                
                # Find name: look backward within y-10 for Chinese text row
                proc_name = ''
                for prev_y in range(y - 10, y):
                    if prev_y in rows:
                        prev_chars = rows[prev_y]
                        prev_sorted = sorted(prev_chars, key=lambda x: x['x0'])
                        prev_text = get_row_text(prev_chars)
                        # Skip if it's a doc number row or digit row
                        if re.match(r'^(\d+[\d\-]*|11232010|SMX)', prev_text):
                            continue
                        # Find Chinese name
                        name_c = [c for c in prev_sorted 
                                  if 60 <= c['x0'] <= 200 and ord(c['text']) > 127 and c['text'] not in (' ', '/')]
                        if name_c:
                            proc_name = ''.join(c['text'] for c in sorted(name_c, key=lambda x: x['x0'])).strip()
                            break
                
                # If no name found in row above, use text from this row
                if not proc_name:
                    row_text = get_row_text(row_chars)
                    # Proc number is at start, name might be elsewhere in row text
                    # but for proc rows like '3' there's no name in same row
                    pass
                
                # Doc number
                row_text = get_row_text(row_chars)
                doc_match = re.search(r'(11232010[\d]+|SMX[\w]+)', row_text)
                doc_no = doc_match.group(1) if doc_match else ''
                
                # ASCO: find first ASCO row after this proc
                after = [ay for ay in page_asco_ys if ay > y]
                if after:
                    nearest_y = min(after, key=lambda ay: ay - y)
                    asco = asco_lookup.get(nearest_y)
                else:
                    before = [ay for ay in page_asco_ys if ay < y]
                    asco = asco_lookup.get(max(before)) if before else None
                
                a_col = asco[0] if asco else ''
                s_col = asco[1] if asco else ''
                c_col = asco[2] if asco else ''
                o_col = asco[3] if asco else ''
                
                # Skip if no proc name found
                if not proc_name:
                    continue
                
                results.append({
                    'processNo': proc_no,
                    'processName': proc_name,
                    'docNo': doc_no,
                    'A': a_col,
                    'S': s_col,
                    'C': c_col,
                    'O': o_col
                })
    
    # Deduplicate by procNo
    seen = {}
    for p in results:
        if p['processNo'] not in seen:
            seen[p['processNo']] = p
    return list(seen.values())


def sort_key(p):
    parts = p['processNo'].replace('-', '.').split('.')
    return [int(x) for x in parts]


def build_excel(processes, output_path, material_code="1907RCP10101"):
    shutil.copy(TEMPLATE_FILE, output_path)
    processes.sort(key=sort_key)
    
    wb = load_workbook(output_path)
    ws = wb['质量计划-工序信息']
    start_row = 10
    
    for i, proc in enumerate(processes):
        r = start_row + i
        seq_no = (i + 1) * 10
        proc_no = proc['processNo']
        proc_name = proc['processName']
        doc_no = proc['docNo']
        col_A = proc['A']
        col_S = proc['S']
        col_C = proc['C']
        col_O = proc['O']
        
        is_need_report = 'Y' if (col_A or col_S or col_C or col_O) else 'N'
        
        # predProcessId
        pred_id = ''
        if '-' in proc_no:
            parent_no = proc_no.split('-')[0]
            for idx, p in enumerate(processes):
                if p['processNo'] == parent_no:
                    pred_id = str((idx + 1) * 10)
                    break
        
        ws.cell(row=r, column=4, value=seq_no)
        ws.cell(row=r, column=6, value=material_code)
        ws.cell(row=r, column=8, value=proc_no)
        ws.cell(row=r, column=9, value=proc_name)
        ws.cell(row=r, column=11, value=col_S if col_S else None)
        ws.cell(row=r, column=12, value='N')
        ws.cell(row=r, column=13, value=doc_no if doc_no else None)
        ws.cell(row=r, column=15, value=None)
        ws.cell(row=r, column=16, value='否')
        ws.cell(row=r, column=17, value=is_need_report)
        ws.cell(row=r, column=18, value=col_A if col_A else None)
        ws.cell(row=r, column=20, value=col_C if col_C else None)
        ws.cell(row=r, column=21, value=col_O if col_O else None)
        ws.cell(row=r, column=23, value=None)
        ws.cell(row=r, column=24, value=pred_id if pred_id else None)
    
    wb.save(output_path)
    return True


def convert_qcp(pdf_path, output_path, qcp_name):
    print(f"\n{'='*60}")
    print(f"转换：{qcp_name}")
    print(f"PDF: {pdf_path}")
    print(f"输出: {output_path}")
    
    processes = parse_qcp(pdf_path)
    processes.sort(key=sort_key)
    
    print(f"\n提取到 {len(processes)} 个工序：")
    for p in processes:
        print(f"  {p['processNo']:8s} | {p['processName'][:20]:20s} | A={p['A'] or '-':3s} S={p['S'] or '-':3s} C={p['C'] or '-':3s} O={p['O'] or '-':3s}")
    
    success = build_excel(processes, output_path)
    
    if success:
        print(f"\n✅ 转换完成：{output_path}")
    else:
        print(f"\n❌ 转换失败")
    
    return success, [p['processNo'] for p in processes]


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    pdf1 = '/home/songsong/.openclaw/media/qqbot/downloads/SMX44400Z62101A04GN Rev.A1 PRE 三门核电项目5、6号机组反应堆冷却剂泵泵壳底脚与泵壳组焊焊接见证件制造质量计划（112320104003171_ _）_1782448358106_ea6200.pdf'
    out1 = '/mnt/d/SynologyDrive/QCP to Excel/SMX44400Z62101A04GN_QCP导入数据.xlsx'
    
    pdf2 = '/home/songsong/.openclaw/media/qqbot/downloads/SMX44400Z63101A04GN Rev.A1 PRE 三门核电项目5、6号机组反应堆冷却剂泵泵壳与安全端焊缝焊接见证件制造质量计划（112320104003172_ _）_1782448384175_6792db.pdf'
    out2 = '/mnt/d/SynologyDrive/QCP to Excel/SMX44400Z63101A04GN Rev.A1 泵壳焊缝见证件QCP.xlsx'
    
    ok1, nums1 = convert_qcp(pdf1, out1, 'Z621底脚组焊')
    ok2, nums2 = convert_qcp(pdf2, out2, 'Z631安全端焊缝')
    
    print(f"\n{'='*60}")
    print("汇总：")
    print(f"Z621: {'✅' if ok1 else '❌'} - {len(nums1)} 个工序")
    print(f"  工序编号: {nums1}")
    print(f"Z631: {'✅' if ok2 else '❌'} - {len(nums2)} 个工序")
    print(f"  工序编号: {nums2}")